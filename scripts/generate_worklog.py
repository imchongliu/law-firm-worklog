"""
Generate a formatted .xlsx worklog from JSON data.

Usage:
  python generate_worklog.py --input data.json --output worklog.xlsx
  python generate_worklog.py --input data.json --output worklog.xlsx --template template.xlsx

Input JSON structure:
{
  "headers": ["日期", "描述", "项目", "案号", "小时"],
  "rows": [
    ["2026-05-03", "审查合同第3条", "项目A", "ABC-2026-001", ""],
    ...
  ],
  "column_widths": {"日期": 12, "描述": 50}   // optional, keyed by header name
}

If --template is provided and is a valid .xlsx file, column widths from the
template's header row are used as defaults (overridden by explicit column_widths).
"""

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Missing openpyxl. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

HEADER_FILL = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
HEADER_FONT = Font(bold=True)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
DEFAULT_COL_WIDTHS = {
    "日期": 12,
    "描述": 50,
    "内容": 50,
    "任务": 50,
    "项目": 18,
    "案号": 22,
    "客户": 18,
    "小时": 10,
    "时长": 10,
}


def read_template_widths(template_path):
    """Read column widths from the first row of a template .xlsx file."""
    from openpyxl import load_workbook

    wb = load_workbook(template_path)
    ws = wb.active
    widths = {}
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        header = ws.cell(row=1, column=col_idx).value
        if header and ws.column_dimensions[letter].width:
            widths[str(header)] = ws.column_dimensions[letter].width
    wb.close()
    return widths


def auto_width(headers, rows, custom_widths=None):
    """Determine column widths: custom > template > heuristics > auto-fit."""
    widths = {}
    for idx, header in enumerate(headers):
        # Check custom widths first
        key = str(header)
        if custom_widths and key in custom_widths:
            widths[idx] = custom_widths[key]
            continue
        # Default heuristic
        widths[idx] = DEFAULT_COL_WIDTHS.get(key, 15)
    return widths


def generate_worklog(headers, rows, output_path, column_widths=None, template_path=None):
    """Write a formatted .xlsx worklog file."""
    wb = Workbook()
    ws = wb.active

    # --- Merge template widths ---
    if template_path:
        try:
            tmpl_widths = read_template_widths(template_path)
            if column_widths:
                tmpl_widths.update(column_widths)
            column_widths = tmpl_widths
        except Exception:
            pass  # template unreadable → fall through to defaults

    widths = auto_width(headers, rows, column_widths)

    # --- Header row ---
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=str(header) if header else "")
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # --- Data rows ---
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value if value is not None else "")
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # --- Column widths ---
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx + 1)].width = width

    # --- Freeze & filter ---
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}1"

    wb.save(output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Generate formatted .xlsx worklog")
    parser.add_argument("--input", required=True, help="Path to input JSON file")
    parser.add_argument("--output", required=True, help="Path to output .xlsx file")
    parser.add_argument("--template", default=None, help="Optional template .xlsx for column widths")
    args = parser.parse_args()

    with open(args.input, "r", encoding="utf-8") as f:
        data = json.load(f)

    headers = data["headers"]
    rows = data["rows"]
    column_widths = data.get("column_widths", None)

    path = generate_worklog(
        headers=headers,
        rows=rows,
        output_path=args.output,
        column_widths=column_widths,
        template_path=args.template,
    )
    print(f"Saved: {path}")


if __name__ == "__main__":
    main()
